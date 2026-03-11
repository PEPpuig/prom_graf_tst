#!/usr/bin/env python3
import requests
import time
import re
import os
from openpyxl import Workbook, load_workbook

METRICS_URL = "http://localhost:8000/metrics"
EXCEL_FILE = "vllm_metrics.xlsx"
INTERVAL = 0.1

# Tiempo de inicio
start_time = time.time()

# Crear archivo Excel si no existe
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Metrics"
    ws.append(['t_rel', 'num_reqs_running', 'num_reqs_waiting', 'kv_cache_perc'])
    wb.save(EXCEL_FILE)

def get_value(lines, metric_substr):
    for line in lines:
        if metric_substr in line:
            match = re.search(r'\{[^}]+\}\s+([\d.e+-]+)', line)
            if match:
                return float(match.group(1))
    return 0.0

while True:
    try:
        resp = requests.get(METRICS_URL, timeout=2)
        lines = resp.text.splitlines()

        running = get_value(lines, 'vllm:num_requests_running')
        waiting = get_value(lines, 'vllm:num_requests_waiting')
        kv_perc = get_value(lines, 'vllm:kv_cache_usage_perc')

        # Tiempo relativo
        t_rel = round(time.time() - start_time, 2)

        # Cargar el archivo Excel y agregar fila
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([
            t_rel,
            round(running, 5),
            round(waiting, 5),
            round(kv_perc, 5)
        ])
        wb.save(EXCEL_FILE)

    except:
        pass

    time.sleep(INTERVAL)
